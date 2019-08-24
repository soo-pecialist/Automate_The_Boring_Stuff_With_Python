#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Aug 23 17:56:33 2019

@author: Soo Hyeon Kim
"""

import openpyxl, sys, os

try:
    filename = sys.argv[1]
except:
    print("You didn't type Excel file name you want to transpose")
    filename = input("what is your file name?: ")

abs_filename = os.path.abspath(filename)

while not os.path.exists(abs_filename):
    filename = input("Your file name is wrong. Please type file name again: ")
    abs_filename = os.path.abspath(filename)

wb = openpyxl.load_workbook(filename)
sheet = wb.active

max_row = sheet.max_row
max_col = sheet.max_column

holder = []
row = []
for row_num in range(1, max_row+1):
    row = []
    
    for col_num in range(1, max_col+1):
        row.append(sheet.cell(row=row_num, column=col_num).value)
    
    holder.append(row)

wb = openpyxl.Workbook()
sheet = wb.active

for row_num in range(1, max_row+1):
    for col_num in range(1, max_col+1):
        sheet.cell(row=col_num, column=row_num).value = \
                    holder[row_num-1][col_num-1]
        
saved_name = filename.strip('.xlsx')+'_inverted'+'.xlsx'
wb.save(saved_name)

print("File '{}' is created!".format(saved_name))

#### Test *****
# !python3 spreadsheet_cell_inverter.py example.xlsx