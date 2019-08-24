#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Aug 23 13:41:55 2019

@author: Soo Hyeon Kim
Your program will look through the spreadsheet, find specific kinds of produce,
and update their prices. 
"""

"""
the prices that you need to update are as follows
Celery      1.19
Garlic      3.07
Lemon       1.27
"""

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# TODO: Loop through the rows and update the prices.
for row_num in range(2, sheet.max_row): # skip the first row
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]
        
wb.save('updated_ProduceSales.xlsx')
print("'updated_ProduceSales.xlsx' is saved!")
    