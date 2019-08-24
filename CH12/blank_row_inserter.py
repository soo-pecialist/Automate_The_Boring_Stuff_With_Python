#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Aug 23 16:48:41 2019

@author: Soo Hyeon Kim
Takes two integers and a filename string from command line. (N, M, filename)
Starting at row N, the program inserts M blank rows into the spread sheet.
e.g., 
!python3 blank_row_inserter.py 3 2 myProduce.xlsx
"""

import sys, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

try:
    n, m, filename = sys.argv[1:4]
except:
    print("You need to input arguments for N, M, and Excel file name!")
    print("By default, N = 3, M = 2, and file name is 'myProduce.xlsx'\n")
    n, m, filename = '3', '2', 'myProduce.xlsx'

print("Start running...")

n = int(n)
m = int(m)

wb = openpyxl.load_workbook(filename)
sheet = wb.active

last_col_letter = get_column_letter(sheet.max_column)
max_row = sheet.max_row
max_col = sheet.max_column

for row_num in range(max_row + m, n-1):   #last row to n's row
    for col_num in range(1, max_col+1):
        sheet.cell(row=row_num, column=col_num).value =\
            sheet.cell(row=row_num - m, column=col_num).value


for row in sheet['A'+str(n) : last_col_letter+str(n+m-1)]:          
    for cell in row:
        cell.value = None
        
saved_name = filename.strip('.xlsx')+'_blanked'+'.xlsx'
wb.save(saved_name)

print("{} blank rows are inserted at row {} of '{}'".format(m, n, filename))
print("File '{}' is created!".format(saved_name))

## Test ***
# !python3 blank_row_inserter.py 5 3 myProduce.xlsx