#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Aug 23 16:21:04 2019

@author: Soo Hyeon Kim
Takes a number N from the command line and creates and NxN multiplication 
table in an Excel spreadsheet

usage:
    python3 multiplication_table.py 6
"""

import openpyxl, sys
from openpyxl.styles import Font

if len(sys.argv) > 1:
    N = int(sys.argv[1])
else: 
    N = int(input("Forgot to input N? Type N please: "))

wb = openpyxl.Workbook()
sheet = wb.active

bold_font = Font(bold=True)

for n in range(1, N+1):
    sheet.cell(row=1, column=n+1).value = n
    sheet.cell(row=1, column=n+1).font = bold_font
    
for n in range(1, N+1):
    sheet.cell(row=n+1, column=1).value = n
    sheet.cell(row=n+1, column=1).font = bold_font

for r in range(2, N+2):   ## n+1 row (first row is header)
    for c in range(2, N+2):  ## n+1 column (first column is header)
        sheet.cell(row=r, column=c).value = (r-1) * (c-1)

wb.save("multiplicatin_table.xlsx")
print("'multiplicatin_table.xlsx' is created for {}x{}!".format(N, N))

## TEST ***
# !python3 multiplication_table.py 15