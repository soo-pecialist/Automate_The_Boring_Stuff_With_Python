#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Aug 26 11:54:22 2019

@author: Soo Hyeon Kim
- Read data from an Excel spreadsheet
- Find all members who have not paid dues for the latest month
- Find their email addresses and send them personalized reminders
"""

import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']

last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value

# TODO: Check each member's payment status
unpaid_members = {}
for r in range(2, sheet.max_row+1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email

# TODO: Log in to email account.
smtp_obj = smtplib('smtp.gmail.com', 587)
smtplib.ehlo()
smtp_obj.starttls()
smtp_obj.login('my_email_address@gmail.com', sys.argv[1])

# TODO: Send out reminder emails.
for name, email in unpaid_members.items():
    body = "Subject: %s dues unpaid. \nDear %s, \n Records show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you!" % (latest_month, name, latest_month)
    print('Sending email to {} ...'.format(email))
    send_mail_status = smtp_obj.sendmail('my_email_address@gmail.com', email, body)
    
    if send_mail_status != {}:
        print('There was a problem sending email to {}: {}'.format(email, send_mail_status))

smtp_obj.quit()


